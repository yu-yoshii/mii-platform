from flask import Flask, render_template, Response, request
from flask_socketio import SocketIO, emit
from datetime import datetime
from pymongo import MongoClient
from bson.objectid import ObjectId
from urllib.parse import quote
import certifi
import base64
import gridfs
import traceback
import cv2
import numpy as np
import io
import csv

# エクセル生成と自動グラフ化用のライブラリ
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import ScatterChart, Series, Reference

app = Flask(__name__)

# ソケットIO定義 (バッファサイズを大きめにして画像・CSV通信を安定化)
socketio = SocketIO(app, cors_allowed_origins="*", max_http_buffer_size=20000000)

# MongoDB接続設定
mongo_uri = "mongodb+srv://e221239_db_user:Y0u667841you@cluster0.1tmhycc.mongodb.net/?appName=Cluster0"
client = MongoClient(mongo_uri, tlsCAFile=certifi.where(), serverSelectionTimeoutMS=20000)
db = client["SNS"]
fs = gridfs.GridFS(db)

@app.route("/")
def index():
    return render_template("index.html")

def get_image_data(image_id):
    try:
        if not fs.exists(image_id):
            return ""
        image_file = fs.get(image_id).read()
        return base64.b64encode(image_file).decode('utf-8')
    except Exception as e:
        print(f"Error: {e}")
        return ""

# =============================================================
# 📥 Socket.IO 通信イベント群
# =============================================================

@socketio.on('load messages')
def load_messages():
    try:
        images = db.images.find().sort('_id', -1).limit(50)
        spectrums = db.spectrums.find().sort('_id', -1).limit(50)
        
        combined_timeline = []
        
        for msg in images:
            img_b64 = get_image_data(msg['image_id'])
            combined_timeline.append({
                'id': str(msg['_id']),
                'date': msg.get('date', datetime.now()),
                'project_name': msg.get('project_name', '未分類'),
                'material_name': msg.get('material_name', '名称不明'),
                'message': msg.get('message', ''),
                'h_val': msg.get('h_val', 'N/A'),
                's_val': msg.get('s_val', 'N/A'),
                'v_val': msg.get('v_val', msg.get('avg_color', 'N/A')),
                'image_data': f"data:image/jpeg;base64,{img_b64}" if img_b64 else "",
                'type': 'image'
            })
            
        for spec in spectrums:
            data_points = len(spec.get('wavelengths', []))
            combined_timeline.append({
                'id': str(spec['_id']),
                'date': spec.get('date', datetime.now()),
                'project_name': spec.get('project_name', '未分類'),
                'material_name': spec.get('material_name', '名称不明'),
                'message': f"【分光データ】{data_points} 点の波長・光度データを保持",
                'h_val': 'N/A', 's_val': 'N/A', 'v_val': 'N/A',
                'image_data': "",
                'type': 'csv',
                # 💡 ここが抜けていたぜッ！過去ログの波長と光度をきっちりJSへ引き渡す！
                'wavelengths': spec.get('wavelengths', []),
                'intensities': spec.get('intensities', [])
            })
            
        combined_timeline.sort(key=lambda x: x['date'])
        
        for item in combined_timeline:
            if isinstance(item['date'], datetime):
                item['date'] = item['date'].strftime('%Y-%m-%d %H:%M:%S')
        
        emit('load all messages', combined_timeline)
        
    except Exception as e:
        traceback.print_exc()

@socketio.on('send message')
def send_message(data):
    try:
        project_name = data.get('project_name', '未分類') 
        material_name = data.get('material_name', '')
        message = data.get('message', '')
        image_data = data.get('image_data', '')

        if image_data and ',' in image_data:
            encoded = image_data.split(',', 1)[1]
            image_bytes = base64.b64decode(encoded)

            nparray = np.frombuffer(image_bytes, np.uint8)
            img = cv2.imdecode(nparray, cv2.IMREAD_COLOR)

            h_val, s_val, v_val = 0, 0, 0
            if img is not None:
                hsv_img = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
                avg_hsv = np.mean(hsv_img, axis=(0, 1))
                h_val = round(float(avg_hsv[0]), 1)
                s_val = round(float(avg_hsv[1]), 1)
                v_val = round(float(avg_hsv[2]), 1)

            image_id = fs.put(image_bytes, filename="upload.jpg")
            result = db.images.insert_one({
                'project_name': project_name,
                'material_name': material_name,
                'image_id': image_id,
                'message': message,
                'h_val': h_val,
                's_val': s_val,
                'v_val': v_val,
                'avg_color': v_val,
                'date': datetime.now()
            })

            # 💡 新しく保存したときも、波長・光度データを即座にグラフに回せるよう配列を追加！
            emit('load one message', {
                'id': str(result.inserted_id),
                'project_name': project_name,
                'material_name': material_name,
                'message': f"【分光データ】{len(wavelengths)} 点の波長・光度データを保持",
                'h_val': 'N/A', 
                's_val': 'N/A', 
                'v_val': 'N/A',
                'image_data': "",
                'type': 'csv',
                'wavelengths': wavelengths,
                'intensities': intensities
            }, broadcast=True)

    except Exception as e:
        traceback.print_exc()

@socketio.on('send csv data')
def handle_csv_data(data):
    try:
        project_name = data.get('project_name', '未分類')
        material_name = data.get('material_name', '名称不明')
        csv_text = data.get('csv_text', '')

        wavelengths = []
        intensities = []

        f = io.StringIO(csv_text)
        reader = csv.reader(f)
        for row in reader:
            if not row or len(row) < 2:
                continue
            try:
                w = float(row[0].strip())
                i = float(row[1].strip())
                wavelengths.append(w)
                intensities.append(i)
            except ValueError:
                continue

        if wavelengths and intensities:
            result = db.spectrums.insert_one({
                'project_name': project_name,
                'material_name': material_name,
                'wavelengths': wavelengths,
                'intensities': intensities,
                'date': datetime.now()
            })
            
            # 💡 新しく保存したときも、波長・光度データを即座にグラフに回せるよう配列を追加！
            emit('load one message', {
                'id': str(result.inserted_id),
                "project_name": project_name,
                "material_name": material_name,
                "message": f"【分光データ】{len(wavelengths)} 点の波長・光度データを保持",
                "h_val": "N/A", "s_val": "N/A", "v_val": "N/A",
                "image_data": "",
                "type": "csv",
                "wavelengths": wavelengths,
                "intensities": intensities
            }, broadcast=True)

    except Exception as e:
        traceback.print_exc()

@socketio.on('delete message')
def delete_message(data):
    try:
        msg_id = data.get('id')
        if not msg_id:
            print("削除エラー: IDが空です")
            return

        target_image = db.images.find_one({"_id": ObjectId(msg_id)})
        if target_image:
            fs.delete(target_image['image_id'])
            db.images.delete_one({"_id": ObjectId(msg_id)})
            print(f"画像データをDBから完全消去 ID: {msg_id}")
            emit('message deleted', {'id': msg_id}, broadcast=True)
            return

        target_spec = db.spectrums.find_one({"_id": ObjectId(msg_id)})
        if target_spec:
            db.spectrums.delete_one({"_id": ObjectId(msg_id)})
            print(f"分光CSVデータをDBから完全消去 ID: {msg_id}")
            emit('message deleted', {'id': msg_id}, broadcast=True)
            return

        print(f"警告: 削除対象のデータがDBに見つかりません ID: {msg_id}")

    except Exception as e:
        print(f"削除失敗: {e}")
        traceback.print_exc()

# =============================================================
# 💾 HTTPルート群（項目別ダウンロード＆エクセル自動グラフ化）
# =============================================================

@app.route("/download_csv")
def download_csv():
    try:
        target_project = request.args.get('project', 'すべて')
        query = {}
        if target_project != 'すべて':
            query['project_name'] = target_project
            
        spectrums = db.spectrums.find(query).sort('date', 1)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['Date', 'Project', 'Material', 'Wavelength (nm)', 'Intensity'])
        
        for spec in spectrums:
            dt = spec.get('date')
            date_str = dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
            p_name = spec.get('project_name', '未分類')
            m_name = spec.get('material_name', '名称不明')
            w_list = spec.get('wavelengths', [])
            i_list = spec.get('intensities', [])
            
            for w, i in zip(w_list, i_list):
                writer.writerow([date_str, p_name, m_name, w, i])
            
        output.seek(0)
        sjis_data = output.getvalue().encode('cp932', errors='replace')
        
        raw_filename = f"spectrum_report_{target_project}.csv" if target_project != 'すべて' else "spectrum_report_all.csv"
        encoded_filename = quote(raw_filename)
        
        return Response(
            sjis_data,
            mimetype="text/csv",
            headers={"Content-disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        traceback.print_exc()
        return "CSV生成エラー", 500

@app.route("/download_excel")
def download_excel():
    try:
        target_project = request.args.get('project', 'すべて')
        query = {}
        if target_project != 'すべて':
            query['project_name'] = target_project

        wb = Workbook()
        
        ws1 = wb.active
        ws1.title = "SpectrumData"
        
        spectrums = db.spectrums.find(query).sort('date', 1)
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws1.append(['Project', 'Material', 'Wavelength (nm)', 'Intensity'])
        for cell in ws1[1]:
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        row_count1 = 1
        series_list = []
        
        for spec in spectrums:
            p_name = spec.get('project_name', '未分類')
            m_name = spec.get('material_name', '名称不明')
            w_list = spec.get('wavelengths', [])
            i_list = spec.get('intensities', [])
            
            start_row = row_count1 + 1
            for w, i in zip(w_list, i_list):
                ws1.append([p_name, m_name, w, i])
                row_count1 += 1
            end_row = row_count1
            
            if start_row <= end_row:
                xvalues = Reference(ws1, min_col=3, min_row=start_row, max_row=end_row)
                yvalues = Reference(ws1, min_col=4, min_row=start_row, max_row=end_row)
                ser = Series(yvalues, xvalues=xvalues, title=f"{m_name} ({p_name})")
                
                ser.graphicalProperties.line.width = 25000
                ser.marker.symbol = "none"
                series_list.append(ser)

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = chr(64 + col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        if series_list:
            chart = ScatterChart()
            chart.title = f"Optical Spectrum ({target_project})"
            chart.style = 13
            chart.x_axis.title = "Wavelength (nm)"
            chart.y_axis.title = "Intensity"
            chart.scatterStyle = "line"
            for ser in series_list:
                chart.series.append(ser)
            ws1.add_chart(chart, "F2")

        ws2 = wb.create_sheet(title="AnalysisData")
        ws2.append(['Date', 'Project', 'Material', 'Hue (色相)', 'Saturation (彩度)', 'Value (明度)', 'Message'])
        for cell in ws2[1]:
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            
        messages = db.images.find(query).sort('date', 1)
        for msg in messages:
            dt = msg.get('date')
            date_str = dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
            ws2.append([
                date_str,
                msg.get('project_name', '未分類'),
                msg.get('material_name', '名称不明'),
                msg.get('h_val', 'N/A'),
                msg.get('s_val', 'N/A'),
                msg.get('v_val', msg.get('avg_color', 'N/A')),
                msg.get('message', '')
            ])

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = chr(64 + col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        raw_filename = f"mi_report_{target_project}.xlsx" if target_project != 'すべて' else "mi_analysis_report.xlsx"
        encoded_filename = quote(raw_filename)
        
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    except Exception as e:
        traceback.print_exc()
        return "Excel生成エラー", 500

if __name__ == "__main__":
    socketio.run(app, debug=True, host="0.0.0.0", port=5000)